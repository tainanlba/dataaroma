#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DATA SCENT × TK美国市场 深度竞品分析 v4b
修复：
  1. 我的产品图片匹配：price_img_map 的 key 改为 +1（1-based），与 _excel_row 对齐
  2. 第四章双榜拆成上下两个独立表格（先销量TOP20，再GMV TOP20）
  3. 第八章：批发价=0 的产品不输出竞争力分析行
  4. 双榜/暴增榜/佣金榜的标题格式统一为：【店铺名】标题
"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8')
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from io import BytesIO
from collections import defaultdict, Counter

BASE = Path(r'D:\鹏工作区域\鹏工作区域\香薰项目')
OUT_DIR = BASE / '产品资料'
TK_DIR  = BASE / 'tk美国搜集翻译转化脚本' / '转化结果'
OUTPUT  = OUT_DIR / 'TK美国市场竞品分析_深度版.xlsx'
TOP_N = 20

print('='*60)
print('DATA SCENT × TK美国市场 深度竞品分析 v4b')
print('='*60)

# ================================================================
# 工具函数
# ================================================================
def parse_num(v):
    if v is None: return None
    s = str(v).strip()
    m = re.match(r'([\d.]+)万', s)
    if m: return float(m.group(1)) * 10000
    try: return float(s.replace(',', ''))
    except: return None

def parse_usd_amount(v):
    if v is None: return None
    s = str(v).strip().replace('$', '').replace(',', '')
    m = re.match(r'([\d.]+)万', s)
    if m: return float(m.group(1)) * 10000
    try: return float(s)
    except: return None

def copy_img_to_cell(src_img, ws_dst, col, row, max_size=120):
    try:
        src_img.ref.seek(0)
        pil_img = PILImage.open(src_img.ref)
        buf = BytesIO()
        pil_img.save(buf, format=pil_img.format or 'PNG')
        buf.seek(0)
        new_img = XLImage(buf)
        scale = min(max_size/new_img.width, max_size/new_img.height, 1.0)
        new_img.width  = int(new_img.width  * scale)
        new_img.height = int(new_img.height * scale)
        new_img.anchor = ws_dst.cell(row=row, column=col).coordinate
        ws_dst.add_image(new_img)
        return True
    except Exception:
        return False

def tb_border(): return Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))

# ================================================================
# 1. 读取我的产品
# ================================================================
print('\n[1/6] 读取 DATA SCENT 产品报价单...')
wb_price   = load_workbook(OUT_DIR / '3_10_香氛机-大客户详细报价单3.xlsx')
ws_price   = wb_price.active

price_img_map = {}
for img in ws_price._images:
    # ★ 修复1：openpyxl 的 _from.row 是 0-based，+1 转为 1-based，与 _excel_row 对齐
    r = img.anchor._from.row + 1
    price_img_map.setdefault(r, []).append(img)

my_products = []
for i, row in enumerate(ws_price.iter_rows(min_row=4, max_row=18)):
    model = str(row[1].value or '').strip()
    if not model or model.startswith('地址') or model == '序号':
        continue
    space    = str(row[3].value or '').strip()
    color    = str(row[4].value or '').strip().replace('\n',' ')
    params   = str(row[5].value or '').strip().replace('\n',' ')
    price_raw = row[6].value
    box      = str(row[7].value or '').strip()
    note     = str(row[8].value or '').strip()
    price_num = 0
    if isinstance(price_raw, (int, float)):
        price_num = int(price_raw)
    elif isinstance(price_raw, str):
        pf = re.findall(r'¥(\d+)', price_raw)
        if pf: price_num = int(pf[0])
    excel_row = i + 4   # 与 price_img_map 的 key 同为 1-based
    my_products.append({
        '型号': model, '覆盖空间': space, '颜色': color,
        '产品参数': params, '批发价格': str(price_raw or ''),
        '批发价格数字': price_num, '箱体参数': box, '备注': note,
        '来源': 'DATA SCENT 产品', '_excel_row': excel_row,
    })

print(f'  ✅ 我的产品: {len(my_products)} 个')
for p in my_products:
    rl = p['批发价格数字']*2   if p['批发价格数字']>0 else 0
    rh = p['批发价格数字']*3.5 if p['批发价格数字']>0 else 0
    print(f"     {p['型号']:<30} 批发¥{p['批发价格数字']:<5} → 零售¥{rl}-{rh}")

# ================================================================
# 2. 读取 TK 数据（含图片缓存）
# ================================================================
print('\n[2/6] 读取 TK 转化结果...')
tk_all = []
seen_titles = set()
tk_img_cache = {}

tk_files = sorted(TK_DIR.glob('*_转化.xlsx'))
for f in tk_files:
    print(f'  读取: {f.name}')
    wb = load_workbook(f)
    ws = wb.active
    fmap = {}
    for img in ws._images:
        r = img.anchor._from.row + 1   # 0-based → 1-based
        fmap.setdefault(r, []).append(img)
    tk_img_cache[str(f)] = fmap

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        title_en = str(row[1] or '').strip()
        if not title_en: continue
        clean = re.sub(r'\s+', ' ', title_en).strip()
        if clean in seen_titles: continue
        seen_titles.add(clean)

        tk_all.append({
            '英文标题': title_en,
            '商品链接': row[2],
            '图片链接': row[3],
            '售价USD': row[4],
            '类目': row[5],
            '评分': row[6],
            '佣金比例': str(row[7] if len(row)>7 else '').strip(),
            '店铺名': str(row[9] if len(row)>9 else '').strip(),
            '店铺销量': row[12] if len(row)>12 else None,
            '国家': row[13] if len(row)>13 else None,
            '达人出单率': row[14] if len(row)>14 else None,
            '近7天销量': row[15] if len(row)>15 else None,
            '近7天销售额': row[16] if len(row)>16 else None,
            '总销量': row[17] if len(row)>17 else None,
            '总销售额': row[18] if len(row)>18 else None,
            '关联达人': row[19] if len(row)>19 else None,
            '中文标题': str(row[21] if len(row)>21 else '').strip(),
            '售价CNY': row[22] if len(row)>22 else '',
            '近7天销售额CNY': row[24] if len(row)>24 else '',
            '总销售额CNY': row[25] if len(row)>25 else '',
            '来源文件': f.stem, '来源': 'TK美国市场',
            '_src_file': str(f), '_src_row': row_idx,
            '售价USD_num':     parse_num(row[4]),
            '总销量_num':      parse_num(row[17]),
            '近7天销量_num':  parse_num(row[15]),
            '总销售额USD':    parse_usd_amount(row[18]),
            '近7天销售额USD': parse_usd_amount(row[16]),
        })

print(f'  ✅ TK竞品数据(去重后): {len(tk_all)} 条')

# ================================================================
# 3. 深度分析
# ================================================================
print('\n[3/6] 执行深度分析...')
prices   = [p for p in tk_all if p['售价USD_num'] and p['售价USD_num']>0]
gmv_all  = [p for p in tk_all if p['总销售额USD'] and p['总销售额USD']>0]
sales_7d = [p for p in tk_all if p['近7天销量_num'] and p['近7天销量_num']>0]

# 3.1 价格带
price_ranges = [
    ('$0-10',   0,  10), ('$10-20',  10,  20), ('$20-30',  20,  30),
    ('$30-50',  30,  50), ('$50-100', 50, 100), ('$100-300',100, 300),
    ('$300+',  300, 99999),
]
price_analysis = []
for label, lo, hi in price_ranges:
    items = [p for p in prices if lo <= p['售价USD_num'] < hi]
    cnt = len(items)
    pct = cnt / len(prices) * 100 if prices else 0
    avgp = sum(p['售价USD_num'] for p in items) / cnt if cnt else 0
    top1 = sorted(items, key=lambda x: x['总销量_num'] or 0, reverse=True)[:1]
    ts = top1[0]['总销量_num'] / 10000 if top1 and top1[0]['总销量_num'] else '-'
    price_analysis.append({'价格带': label, '数量': cnt, '占比%': round(pct,1),
                          '平均售价': round(avgp,2), 'TOP1销量(万)': ts})

# 3.2 类目
cat_counter = Counter()
cat_gmv = defaultdict(float)
cat_sales = defaultdict(float)
for p in tk_all:
    cat = str(p['类目'] or '未知').strip()
    cat_counter[cat] += 1
    if p['总销售额USD']: cat_gmv[cat] += p['总销售额USD']
    if p['总销量_num']:    cat_sales[cat] += p['总销量_num']
cat_analysis = []
for cat, cnt in cat_counter.most_common(20):
    cat_analysis.append({
        '类目': cat, '商品数': cnt,
        '占比%': round(cnt/len(tk_all)*100,1),
        '总GMV(万美元)': round(cat_gmv[cat]/10000,1) if cat_gmv[cat] else 0,
        '平均销量(万)': round(cat_sales[cat]/cnt/10000,1) if cnt else 0,
    })

# 3.3 高佣金
commission_items = []
for p in tk_all:
    m = re.search(r'(\d+)%', str(p.get('佣金比例','')))
    if m:
        commission_items.append((int(m.group(1)), p))
commission_items.sort(key=lambda x: x[0], reverse=True)
top_commission = commission_items[:TOP_N]

# 3.4 近7天暴增榜
trending = []
for p in tk_all:
    if p['近7天销量_num'] and p['总销量_num'] and p['总销量_num'] > 0:
        ratio = p['近7天销量_num'] / p['总销量_num']
        if ratio > 0.05:
            trending.append((ratio, p))
trending.sort(key=lambda x: x[0], reverse=True)
trending_top = trending[:TOP_N]

# 3.5 店铺
shop_counter = Counter()
shop_gmv = defaultdict(float)
for p in tk_all:
    shop = str(p.get('店铺名') or '').strip()
    if shop:
        shop_counter[shop] += 1
        if p['总销售额USD']: shop_gmv[shop] += p['总销售额USD']
top_shops = []
for shop, cnt in shop_counter.most_common(TOP_N):
    top_shops.append({'店铺名': shop, '商品数': cnt,
                     '总GMV(万美元)': round(shop_gmv[shop]/10000,1) if shop_gmv[shop] else 0})

# 3.6 国家
country_counter = Counter()
for p in tk_all:
    c = str(p.get('国家') or '未知').strip()
    country_counter[c] += 1
country_analysis = country_counter.most_common(15)

# 3.7 TOP 双榜
top_sales = sorted([p for p in tk_all if p['总销量_num']],
                   key=lambda x: x['总销量_num'], reverse=True)[:TOP_N]
top_gmv   = sorted([p for p in tk_all if p['总销售额USD']],
                   key=lambda x: x['总销售额USD'], reverse=True)[:TOP_N]

# 3.8 与我产品相关竞品（按价格带匹配）
related_competitors = []
for mp in my_products:
    if mp['批发价格数字'] <= 0:
        continue
    lo = mp['批发价格数字'] * 2 / 7 * 0.8
    hi = mp['批发价格数字'] * 3.5 / 7 * 1.2
    comps = [p for p in tk_all if p['售价USD_num'] and lo <= p['售价USD_num'] <= hi]
    comps.sort(key=lambda x: (x['总销量_num'] or 0), reverse=True)
    related_competitors.append((mp, comps[:5]))
seen = set()
related_top = []
for mp, comps in related_competitors:
    for p in comps:
        key = p['英文标题'][:50]
        if key not in seen:
            seen.add(key)
            related_top.append((mp, p))
            if len(related_top) >= TOP_N:
                break
    if len(related_top) >= TOP_N:
        break

# 3.9 价格竞争力评分
for mp in my_products:
    if mp['批发价格数字'] > 0:
        rl = mp['批发价格数字'] * 2 / 7
        rh = mp['批发价格数字'] * 3.5 / 7
        same_range = [p for p in prices if rl*0.7 <= p['售价USD_num'] <= rh*1.3]
        if same_range:
            avg_comp = sum(p['售价USD_num'] for p in same_range) / len(same_range)
            mp['竞争力评分'] = '高' if rl < avg_comp * 0.85 else ('中' if rl < avg_comp * 1.1 else '低')
            mp['竞品均价'] = round(avg_comp, 1)
        else:
            mp['竞争力评分'] = '未知'
            mp['竞品均价'] = 0
    else:
        mp['竞争力评分'] = ''
        mp['竞品均价'] = 0

# 3.10 市场空白点
blank_spots = []
for pa in price_analysis:
    pa_lo = float(re.sub(r'[^\d.]', '', pa['价格带'].split('-')[0])) if '-' in pa['价格带'] else 0
    pa_hi = float(re.sub(r'[^\d.]', '', pa['价格带'].split('-')[1])) if '-' in pa['价格带'] else 99999
    for ca in cat_analysis[:5]:
        cnt = len([p for p in tk_all if p['售价USD_num'] and
                    pa_lo <= p['售价USD_num'] < pa_hi and
                    str(p['类目'] or '').strip() == ca['类目']])
        if cnt < 20:
            blank_spots.append((pa['价格带'], ca['类目'], cnt))
blank_spots.sort(key=lambda x: x[2])

print(f'  ✅ 分析完成: 价格带{len(price_analysis)}档 | 类目{len(cat_analysis)}种 | 店铺{len(top_shops)}个')
print(f'  ✅ 相关竞品: {len(related_top)} 个 | 空白点: {len(blank_spots)} 个')

# ================================================================
# 4. 写入 Excel
# ================================================================
print('\n[4/6] 写入 Excel...')
wb_out = Workbook()

# ---------- 样式快捷 ----------
HDR_BG   = '2F5496'
RED_BG   = 'C00000'
INSIGHT  = 'C00000'
TITLE_BG = '1F4E79'
BAND_BG1 = 'D9E1F2'   # 浅蓝表头
BAND_BG2 = 'E9EEF7'   # 更浅蓝行头

def wc(ws, r, c, v, bold=False, color='000000', size=11, bg=None, align='center', wrap=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, color=color, size=size)
    if bg: cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border = tb_border()
    return cell

# ================================================================
# Sheet1: 我的产品
# ================================================================
ws1 = wb_out.active
ws1.title = '我的产品'
hd1 = ['序号','型号','覆盖空间','颜色','产品参数','批发价格(¥)','箱体参数','备注','来源','竞争力评分','竞品均价(USD)','产品图片']
for c, h in enumerate(hd1, 1):
    wc(ws1, 1, c, h, bold=True, bg=HDR_BG, color='FFFFFF')
ws1.row_dimensions[1].height = 25

for i, p in enumerate(my_products, 2):
    wc(ws1, i, 1,  i-1,                        align='center')
    wc(ws1, i, 2,  p['型号'],                    align='center')
    wc(ws1, i, 3,  p['覆盖空间'],               align='center')
    wc(ws1, i, 4,  p['颜色'],                    align='left')
    wc(ws1, i, 5,  p['产品参数'],               align='left')
    wc(ws1, i, 6,  p['批发价格'],               align='center')
    wc(ws1, i, 7,  p['箱体参数'],               align='left')
    wc(ws1, i, 8,  p['备注'],                    align='left')
    wc(ws1, i, 9,  p['来源'],                    align='center')
    score = p.get('竞争力评分', '')
    sc = {'高':'375623','中':'BF8F00','低':'C00000'}.get(score, '000000')
    wc(ws1, i, 10, score, bold=True, color=sc,   align='center')
    wc(ws1, i, 11, p.get('竞品均价', 0) or '',  align='center')
    # ★ 修复1：现在 _excel_row 与 price_img_map 的 key 同为 1-based
    er = p['_excel_row']
    if er and er in price_img_map:
        ok = copy_img_to_cell(price_img_map[er][0], ws1, len(hd1), i)
        if ok:
            ws1.row_dimensions[i].height = 90
    ws1.row_dimensions[i].height = max(ws1.row_dimensions[i].height or 15, 40)

ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['E'].width = 50
ws1.column_dimensions['G'].width = 30
ws1.column_dimensions[get_column_letter(len(hd1))].width = 20
print('  ✅ Sheet[我的产品] 完成')

# ================================================================
# Sheet2: TK竞品
# ================================================================
ws2 = wb_out.create_sheet('TK美国市场竞品')
hd2 = ['序号','店铺名','英文标题','中文标题','售价(USD)','售价(¥)','类目','评分','佣金比例',
        '店铺销量','总销量','近7天销量','总销售额(USD)','近7天销售额(USD)',
        '总销售额(¥)','近7天销售额(¥)','关联达人','商品链接','来源文件','商品图片']
for c, h in enumerate(hd2, 1):
    wc(ws2, 1, c, h, bold=True, bg=RED_BG, color='FFFFFF')
ws2.row_dimensions[1].height = 25

img_col = len(hd2)
for i, item in enumerate(tk_all, 2):
    wc(ws2, i, 1,  i-1,                                    align='center')
    wc(ws2, i, 2,  item.get('店铺名','') or '',           align='center')
    wc(ws2, i, 3,  item['英文标题'],                       align='left')
    wc(ws2, i, 4,  item.get('中文标题','') or '',          align='left')
    wc(ws2, i, 5,  item['售价USD'],                        align='center')
    wc(ws2, i, 6,  item.get('售价CNY','') or '',          align='center')
    wc(ws2, i, 7,  item.get('类目','') or '',             align='center')
    wc(ws2, i, 8,  item.get('评分','') or '',             align='center')
    wc(ws2, i, 9,  item.get('佣金比例','') or '',         align='center')
    wc(ws2, i, 10, item.get('店铺销量','') or '',         align='center')
    wc(ws2, i, 11, item.get('总销量','') or '',           align='center')
    wc(ws2, i, 12, item.get('近7天销量','') or '',       align='center')
    wc(ws2, i, 13, item.get('总销售额','') or '',         align='center')
    wc(ws2, i, 14, item.get('近7天销售额','') or '',     align='center')
    wc(ws2, i, 15, item.get('总销售额CNY','') or '',     align='center')
    wc(ws2, i, 16, item.get('近7天销售额CNY','') or '',  align='center')
    wc(ws2, i, 17, item.get('关联达人','') or '',         align='center')
    wc(ws2, i, 18, item.get('商品链接','') or '',         align='left')
    wc(ws2, i, 19, item.get('来源文件','') or '',         align='center')
    # 嵌入图片
    sf = item.get('_src_file')
    sr = item.get('_src_row')
    if sf and sr and sf in tk_img_cache and sr in tk_img_cache[sf]:
        ok = copy_img_to_cell(tk_img_cache[sf][sr][0], ws2, img_col, i)
        if ok:
            ws2.row_dimensions[i].height = 90
    ws2.row_dimensions[i].height = max(ws2.row_dimensions[i].height or 15, 40)
    if i % 100 == 0:
        print(f'  TK竞品写入: {i-1}/{len(tk_all)}')

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['C'].width = 45
ws2.column_dimensions['D'].width = 35
ws2.column_dimensions[get_column_letter(img_col)].width = 20
print(f'  ✅ Sheet[TK美国市场竞品] 完成，{len(tk_all)}行含图')

# ================================================================
# Sheet3: 深度分析报告
# ================================================================
print('\n[5/6] 写入深度分析报告...')
ws3 = wb_out.create_sheet('深度分析报告')
ws3.sheet_view.showGridLines = False

def w3(r, c, v, bold=False, color='000000', bg=None, align='center', size=11, merge_to=None):
    """ws3 专用写入，可选合并单元格"""
    cell = wc(ws3, r, c, v, bold=bold, color=color, size=size, bg=bg, align=align)
    if merge_to:
        ws3.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge_to)
    return r

def section_title(r, text):
    w3(r, 1, text, bold=True, size=14, bg=TITLE_BG, color='FFFFFF', merge_to=9)
    ws3.row_dimensions[r].height = 30
    return r + 2

def insight_row(r, text):
    w3(r, 1, text, bold=True, color=INSIGHT, merge_to=9)
    return r + 2

def table_header(r, cols, bg=BAND_BG1):
    for c, h in enumerate(cols, 1):
        w3(r, c, h, bold=True, bg=bg, color='000000')
    ws3.row_dimensions[r].height = 20
    return r + 1

def data_row(r, cols, align_list=None):
    for c, v in enumerate(cols, 1):
        al = (align_list or [])[c-1] if align_list else 'center'
        w3(r, c, v, align=al)
    ws3.row_dimensions[r].height = 40
    return r + 1

# -------- 报告正文 --------
r = 1

# 一、市场概况
r = section_title(r, 'DATA SCENT × TK美国市场 深度竞品分析报告')
r = section_title(r, '一、市场概况')
w3(r, 1, f'有效竞品数量: {len(tk_all)} 条（去重后）'); r += 1
w3(r, 1, f"数据来源: {len(tk_files)} 个关键词（aroma diffuser / scent diffuser / scent machine / waterless diffuser / 香薰机）"); r += 1
min_p = min((p['售价USD_num'] for p in prices), default=0)
max_p = max((p['售价USD_num'] for p in prices), default=0)
w3(r, 1, f'竞品价格区间: ${min_p:.2f} - ${max_p:.2f}'); r += 1
avg_price = sum(p['售价USD_num'] for p in prices)/len(prices) if prices else 0
w3(r, 1, f'竞品平均售价: ${avg_price:.2f}'); r += 1
total_gmv = sum(p['总销售额USD'] for p in gmv_all)/10000 if gmv_all else 0
w3(r, 1, f'竞品总GMV（估算）: ${total_gmv:.1f} 万美元'); r += 1
avg7 = sum(p['近7天销量_num'] for p in sales_7d)/len(sales_7d) if sales_7d else 0
w3(r, 1, f'近7天平均销量: {avg7:.0f} 件/商品'); r += 2

# 二、价格带
r = section_title(r, '二、价格带分布分析')
r = table_header(r, ['价格带','商品数','占比%','平均售价(USD)','TOP1销量(万)'])
for pa in price_analysis:
    r = data_row(r, [pa['价格带'], pa['数量'], f"{pa['占比%']}%",
                      f"${pa['平均售价']:.2f}", pa['TOP1销量(万)']])
r += 1
r = insight_row(r, '【洞察】$10-20是最大价格带(32.8%)，但$30-50价格带竞争较少(13.4%)，是品质差异化机会窗口')
r += 2

# 三、类目 TOP20
r = section_title(r, '三、类目分布分析（TOP20）')
r = table_header(r, ['类目','商品数','占比%','总GMV(万美元)','平均销量(万)'])
for ca in cat_analysis[:TOP_N]:
    name = ca['类目'][:36] if len(ca['类目'])>36 else ca['类目']
    r = data_row(r, [name, ca['商品数'], f"{ca['占比%']}%",
                      f"${ca['总GMV(万美元)']:.1f}", f"{ca['平均销量(万)']:.1f}"])
r += 1
r = insight_row(r, '【洞察】车载香薰和家用香薰是主流类目，无水/冷雾扩香器在高端市场($50+)有溢价空间')
r += 2

# 四、销量TOP20（独立表格）
r = section_title(r, '四、销量 TOP20')
r = table_header(r, ['排名','店铺名','商品标题','售价(USD)','总销量','图片'])
for i, p in enumerate(top_sales, 1):
    shop  = (p.get('店铺名') or '')[:18]
    title = ((p.get('中文标题') or p['英文标题']) or '')[:40]
    r = data_row(r, [f'{i}.', shop, title,
                        f"${p['售价USD']}" if p.get('售价USD') else '-',
                        p.get('总销量','-') or '-', ''])
    # 嵌入图片
    sf = p.get('_src_file'); sr = p.get('_src_row')
    if sf and sr and sf in tk_img_cache and sr in tk_img_cache[sf]:
        copy_img_to_cell(tk_img_cache[sf][sr][0], ws3, 6, r-1, max_size=80)
    ws3.row_dimensions[r-1].height = 70
r += 2

# 四（续）：GMV TOP20（独立表格）
r = section_title(r, '四（续）、销售额(GMV) TOP20')
r = table_header(r, ['排名','店铺名','商品标题','售价(USD)','总销售额(USD)','图片'])
for i, p in enumerate(top_gmv, 1):
    shop  = (p.get('店铺名') or '')[:18]
    title = ((p.get('中文标题') or p['英文标题']) or '')[:40]
    gmv = f"${p['总销售额USD']/10000:.1f}万" if p.get('总销售额USD') else '-'
    r = data_row(r, [f'{i}.', shop, title,
                        f"${p['售价USD']}" if p.get('售价USD') else '-', gmv, ''])
    sf = p.get('_src_file'); sr = p.get('_src_row')
    if sf and sr and sf in tk_img_cache and sr in tk_img_cache[sf]:
        copy_img_to_cell(tk_img_cache[sf][sr][0], ws3, 6, r-1, max_size=80)
    ws3.row_dimensions[r-1].height = 70
r += 1
r = insight_row(r, '【洞察】销量TOP1和GMV TOP1不同——低价走量 vs 高价走利润是两种打法，需根据供应链优势选择')
r += 2

# 五、近7天暴增榜
r = section_title(r, '五、近7天销量暴增榜（趋势机会）')
r = table_header(r, ['排名','店铺名','商品标题','售价(USD)','近7天销量','总销量','7天占比','图片'])
for i, (ratio, p) in enumerate(trending_top, 1):
    shop  = (p.get('店铺名') or '')[:15]
    title = ((p.get('中文标题') or p['英文标题']) or '')[:35]
    r = data_row(r, [f'{i}.', shop, title,
                        f"${p['售价USD']}" if p.get('售价USD') else '-',
                        p.get('近7天销量','-') or '-',
                        p.get('总销量','-') or '-',
                        f'{ratio*100:.1f}%', ''])
    sf = p.get('_src_file'); sr = p.get('_src_row')
    if sf and sr and sf in tk_img_cache and sr in tk_img_cache[sf]:
        copy_img_to_cell(tk_img_cache[sf][sr][0], ws3, 8, r-1, max_size=80)
    ws3.row_dimensions[r-1].height = 70
r += 1
r = insight_row(r, '【洞察】近7天销量占比>5%的商品是正在爆发的趋势品，可快速跟进类似产品/卖点')
r += 2

# 六、高佣金爆品
r = section_title(r, '六、高佣金爆品分析（达人带货机会）')
r = table_header(r, ['排名','店铺名','商品标题','佣金','售价(USD)','总销量','图片'])
for i, (comm_pct, p) in enumerate(top_commission, 1):
    shop  = (p.get('店铺名') or '')[:15]
    title = ((p.get('中文标题') or p['英文标题']) or '')[:35]
    r = data_row(r, [f'{i}.', shop, title, f'{comm_pct}%',
                        f"${p['售价USD']}" if p.get('售价USD') else '-',
                        p.get('总销量','-') or '-', ''])
    sf = p.get('_src_file'); sr = p.get('_src_row')
    if sf and sr and sf in tk_img_cache and sr in tk_img_cache[sf]:
        copy_img_to_cell(tk_img_cache[sf][sr][0], ws3, 7, r-1, max_size=80)
    ws3.row_dimensions[r-1].height = 70
r += 1
r = insight_row(r, '【洞察】佣金>15%且销量高的商品，说明达人愿意带——可以同样找达人推我们的同类产品')
r += 2

# 七、TOP店铺
r = section_title(r, '七、TOP店铺分析（竞争对手监控）')
r = table_header(r, ['排名','店铺名','商品数','总GMV(万美元)','平均GMV/商品(USD)'])
for i, si in enumerate(top_shops, 1):
    avg = (si['总GMV(万美元)']*10000) / si['商品数'] if si['商品数'] else 0
    r = data_row(r, [f'{i}.', si['店铺名'][:25], si['商品数'],
                      f"${si['总GMV(万美元)']:.1f}万", f"${avg:.0f}"])
r += 2

# 八、DATA SCENT 精准定位（★ 修复3：只输出有批发价的产品）
r = section_title(r, '八、DATA SCENT 精准定位分析')
r = table_header(r, ['型号','批发价(¥)','建议零售价(USD)','竞品均价(USD)','竞争力','对标竞品数'])
for p in my_products:
    if p['批发价格数字'] <= 0:
        continue   # ★ 修复3：无批发价不输出
    rl = p['批发价格数字']*2/7
    rh = p['批发价格数字']*3.5/7
    comps = [x for x in prices if rl*0.8 <= x['售价USD_num'] <= rh*1.2]
    score = p.get('竞争力评分','')
    sc = {'高':'375623','中':'BF8F00','低':'C00000'}.get(score, '000000')
    r = data_row(r, [p['型号'], f"¥{p['批发价格数字']}",
                      f"${rl:.0f}-{rh:.0f}",
                      f"${p.get('竞品均价',0):.1f}" if p.get('竞品均价',0)>0 else '-',
                      score, f"{len(comps)}个"])
    ws3.cell(row=r-1, column=5).font = Font(bold=True, color=sc)
r += 1
r = insight_row(r, '【建议】TC100(批发¥45)对标$13-22竞争激烈；LD500(批发¥99)对标$28-50竞争少毛利高——优先打LD500')
r += 2

# 八（续）：与我产品相关的前10竞品深度对比
r = section_title(r, '八（续）、与我产品相关的前10竞品深度对比')
r = table_header(r, ['序号','对标产品','竞品店铺','竞品标题','竞品售价(USD)','竞品总销量','竞品评分','佣金','图片'])
for i, (mp, p) in enumerate(related_top, 1):
    r = data_row(r, [i, mp['型号'],
                        (p.get('店铺名') or '')[:15],
                        ((p.get('中文标题') or p['英文标题']) or '')[:40],
                        p.get('售价USD','-') or '-',
                        p.get('总销量','-') or '-',
                        p.get('评分','-') or '-',
                        p.get('佣金比例','-') or '-', ''])
    sf = p.get('_src_file'); sr = p.get('_src_row')
    if sf and sr and sf in tk_img_cache and sr in tk_img_cache[sf]:
        copy_img_to_cell(tk_img_cache[sf][sr][0], ws3, 9, r-1, max_size=80)
    ws3.row_dimensions[r-1].height = 80
r += 2

# 九、国家分布
r = section_title(r, '九、国家分布（市场机会）')
for country, cnt in country_analysis:
    pct = cnt/len(tk_all)*100
    w3(r, 1, f'  {country}: {cnt} 个商品 ({pct:.1f}%)'); r += 1
r += 2

# 十、市场空白点
r = section_title(r, '十、市场空白点分析（竞争少的机会区）')
r = table_header(r, ['价格带','类目','该组合竞品数','机会评级'])
for (price_label, cat, cnt) in blank_spots[:15]:
    level = '★★★★★ 强烈推荐' if cnt < 10 else ('★★★☆☆ 可考虑' if cnt < 20 else '★★☆☆☆ 一般')
    r = data_row(r, [price_label, cat[:30], cnt, level])
r += 2

# 十一、90天行动路线图
r = section_title(r, '十一、90天行动路线图')
actions = [
    ('第1-15天',  '确定TC100/LD500零售定价，对标TK TOP20竞品调整卖点文案和主图风格'),
    ('第16-30天', '完成TC100/LD500样品拍摄（主图+详情页+短视频），准备TK Shop上架素材'),
    ('第31-50天', '找10-20个家居/车载垂类达人（粉丝5-50万），寄样+佣金10-15%，测第一批内容'),
    ('第51-70天', 'TK Shop正式上架，同步跑TK Ads($50/天测试），盯7天数据快速迭代'),
    ('第71-90天', '复盘数据，决定是否扩品类(TX500/TJ2500)，制定Q2-Q3增长计划'),
]
for period, action in actions:
    w3(r, 1, period, bold=True, bg=BAND_BG2, align='center')
    w3(r, 2, action, align='left', merge_to=9)
    r += 1
r += 1

# 十二、风险提示
r = section_title(r, '十二、风险提示')
risks = [
    'C端品牌认知为0，前期需大量内容/达人投入建立信任',
    '物流成本：大空间香薰机(TJ2500/TS8000)体积大，海外仓成本需核算',
    '售后风险：精油/雾化片耗材属性，需建立海外售后/补寄体系',
    'TK政策风险：电商政策变动频繁，需同步布局独立站/亚马逊多渠道',
]
for risk in risks:
    w3(r, 1, '⚠ ' + risk, color=INSIGHT, merge_to=9)
    r += 1

# 设置列宽
ws3.column_dimensions['A'].width = 22
for col_let in ['B','C','D','E','F','G','H','I']:
    ws3.column_dimensions[col_let].width = 18
ws3.column_dimensions['C'].width = 35

# ================================================================
# 6. 保存
# ================================================================
print('\n[6/6] 保存文件...')
wb_out.save(OUTPUT)
print(f'\n✅ 完成！保存至: {OUTPUT}')
print(f'文件大小: {OUTPUT.stat().st_size//1024} KB')
print(f'Sheets: 我的产品 / TK美国市场竞品 / 深度分析报告')
print(f'\n核心数据预览：')
print(f'  竞品总数: {len(tk_all)}')
print(f'  总GMV: ${total_gmv:.1f}万')
print(f'  $30-50价格带竞品: {price_analysis[3]["数量"]}个({price_analysis[3]["占比%"]}%)')
print(f'  近7天暴增商品: {len(trending)}个')
print(f'  高佣金(>15%)商品: {len([c for c,p in commission_items if c>=15])}个')
print(f'  与我产品相关竞品: {len(related_top)}个')
print(f'  市场空白点: {len(blank_spots)}个')
