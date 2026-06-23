#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook(r'D:\鹏工作区域\鹏工作区域\香薰项目\产品资料\TK美国市场竞品分析_深度版.xlsx')

print('=== 工作表 ===')
for name in wb.sheetnames:
    ws = wb[name]
    print(f'  {name}: {ws.max_row}行 x {ws.max_column}列')

ws3 = wb['深度分析报告']
print()
print('=== 深度分析报告 章节结构 ===')
for r in range(1, ws3.max_row+1):
    v = ws3.cell(r, 1).value
    if v and ('、' in str(v) or ')' in str(v)):
        print(f'  R{r}: {str(v)[:60]}')

print()
for name in ['我的产品', 'TK美国市场竞品', '深度分析报告']:
    ws = wb[name]
    print(f'{name} 图片数: {len(ws._images)}')

wb.close()
